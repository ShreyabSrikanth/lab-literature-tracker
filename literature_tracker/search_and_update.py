"""
Literature Tracker — Auto-search and update spreadsheet
Searches PubMed for new CGM+dietary and multi-omics datasets,
then adds any new findings to the Excel tables.

Run manually:  python search_and_update.py
Run on schedule: GitHub Actions (see .github/workflows/weekly_update.yml)
"""

import json
import time
import datetime
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ─── CONFIG ──────────────────────────────────────────────────────────────────

SPREADSHEET_PATH = "data/literature_tables.xlsx"
LOG_PATH         = "data/search_log.json"

# PubMed searches for Table 1 (CGM + dietary datasets)
CGM_QUERIES = [
    "CGM dietary dataset open access meal times",
    "continuous glucose monitoring food log publicly available",
    "CGM nutrition dataset open source participants",
    "glucose monitoring dietary record download",
]

# PubMed searches for Table 2 (multi-omics acute time series)
OMICS_QUERIES = [
    "multi-omics acute time series human open access perturbation",
    "metabolomics proteomics transcriptomics exercise acute human dataset",
    "postprandial omics time series human publicly available",
    "omics short-term response human perturbation open data",
    "systems vaccinology multi-omics time series human open",
]

PUBMED_BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"

# ─── PUBMED SEARCH ───────────────────────────────────────────────────────────

def search_pubmed(query: str, max_results: int = 10) -> list[dict]:
    """
    Search PubMed with a query string.
    Returns a list of dicts with title, authors, year, doi, pmid, abstract.
    """
    # Step 1: search for IDs
    search_url = f"{PUBMED_BASE}/esearch.fcgi"
    search_params = {
        "db": "pubmed",
        "term": query,
        "retmax": max_results,
        "retmode": "json",
        "sort": "pub+date",   # newest first
        "datetype": "pdat",
        "reldate": 365,       # last 12 months only
    }

    try:
        r = requests.get(search_url, params=search_params, timeout=15)
        r.raise_for_status()
        ids = r.json().get("esearchresult", {}).get("idlist", [])
    except Exception as e:
        print(f"  [search error] {query}: {e}")
        return []

    if not ids:
        return []

    time.sleep(0.5)  # be polite to PubMed API

    # Step 2: fetch details for those IDs
    fetch_url = f"{PUBMED_BASE}/efetch.fcgi"
    fetch_params = {
        "db": "pubmed",
        "id": ",".join(ids),
        "retmode": "xml",
        "rettype": "abstract",
    }

    try:
        r = requests.get(fetch_url, params=fetch_params, timeout=15)
        r.raise_for_status()
    except Exception as e:
        print(f"  [fetch error] {e}")
        return []

    # Parse XML manually (avoids needing lxml/BeautifulSoup)
    results = parse_pubmed_xml(r.text, ids)
    time.sleep(0.5)
    return results


def parse_pubmed_xml(xml_text: str, ids: list[str]) -> list[dict]:
    """
    Very lightweight XML parser — extracts title, year, DOI, abstract.
    Avoids external dependencies (no lxml or bs4 needed).
    """
    import xml.etree.ElementTree as ET
    results = []
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return []

    for article in root.findall(".//PubmedArticle"):
        try:
            pmid = article.findtext(".//PMID", "")
            title = article.findtext(".//ArticleTitle", "No title")
            year  = article.findtext(".//PubDate/Year", "")
            if not year:
                year = article.findtext(".//PubDate/MedlineDate", "")[:4]

            # Authors
            authors = []
            for author in article.findall(".//Author")[:3]:
                last  = author.findtext("LastName", "")
                first = author.findtext("ForeName", "")
                if last:
                    authors.append(f"{last} {first[0]}." if first else last)
            author_str = ", ".join(authors) + (" et al." if len(authors) == 3 else "")

            # DOI
            doi = ""
            for id_elem in article.findall(".//ArticleId"):
                if id_elem.get("IdType") == "doi":
                    doi = id_elem.text or ""

            # Abstract (first 400 chars)
            abstract_parts = article.findall(".//AbstractText")
            abstract = " ".join(p.text or "" for p in abstract_parts)[:400]

            # Journal
            journal = article.findtext(".//Journal/Title", "")

            results.append({
                "pmid":     pmid,
                "title":    title,
                "authors":  author_str,
                "year":     year,
                "doi":      doi,
                "journal":  journal,
                "abstract": abstract,
            })
        except Exception:
            continue

    return results


# ─── RELEVANCE FILTERING ─────────────────────────────────────────────────────

# Keywords that strongly suggest a paper describes an actual dataset
DATASET_SIGNALS = [
    "dataset", "data set", "publicly available", "open access",
    "freely available", "data sharing", "repository", "download",
    "participants", "cohort", "open-source", "figshare", "zenodo",
    "physionet", "metabolomics workbench", "metaboLights", "GEO",
    "deposited", "data availability",
]

# Keywords that suggest it's NOT a dataset paper (methods, reviews, clinical trials)
NOISE_SIGNALS = [
    "review", "meta-analysis", "systematic review", "clinical trial",
    "randomized", "algorithm", "deep learning", "machine learning model",
    "prediction model", "case report",
]

def is_relevant(paper: dict, table: str) -> tuple[bool, str]:
    """
    Returns (is_relevant, reason).
    Filters out papers that are clearly not new dataset papers.
    """
    text = (paper["title"] + " " + paper["abstract"]).lower()

    # Must mention actual data availability
    has_dataset_signal = any(kw in text for kw in DATASET_SIGNALS)
    if not has_dataset_signal:
        return False, "no dataset availability signal"

    # Skip obvious noise
    is_noise = any(kw in text for kw in NOISE_SIGNALS)
    if is_noise:
        return False, "appears to be a review/model paper"

    # Table-specific checks
    if table == "cgm":
        needs = ["cgm", "continuous glucose", "glucose monitor"]
        diet  = ["diet", "meal", "food", "nutrition", "dietary"]
        if not any(k in text for k in needs):
            return False, "no CGM mention"
        if not any(k in text for k in diet):
            return False, "no dietary data mention"

    elif table == "omics":
        omics_terms = [
            "metabolom", "proteom", "transcriptom", "genomic",
            "multi-omic", "lipidom", "epigenom",
        ]
        acute_terms = [
            "acute", "time series", "time-series", "time point",
            "longitudinal", "minutes", "hours", "postprandial",
            "exercise", "perturbation", "challenge",
        ]
        if not any(k in text for k in omics_terms):
            return False, "no omics mention"
        if not any(k in text for k in acute_terms):
            return False, "no acute/time-series mention"

    return True, "passes filters"


# ─── SPREADSHEET MANAGEMENT ──────────────────────────────────────────────────

def load_existing_pmids() -> set[str]:
    """Load PMIDs already in the spreadsheet to avoid duplicates."""
    log_file = Path(LOG_PATH)
    if not log_file.exists():
        return set()
    with open(log_file) as f:
        data = json.load(f)
    return set(data.get("seen_pmids", []))


def save_seen_pmids(pmids: set[str]):
    """Save all seen PMIDs to the log file."""
    log_file = Path(LOG_PATH)
    log_file.parent.mkdir(parents=True, exist_ok=True)
    existing = {}
    if log_file.exists():
        with open(log_file) as f:
            existing = json.load(f)
    existing["seen_pmids"] = list(pmids)
    existing["last_run"] = datetime.datetime.now().isoformat()
    with open(log_file, "w") as f:
        json.dump(existing, f, indent=2)


def style_new_row(ws, row_num: int, num_cols: int):
    """Apply light green highlight to newly added rows so they're easy to spot."""
    thin = Side(style="thin", color="AAAAAA")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = PatternFill("solid", start_color="E2EFDA")  # light green
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border    = bdr


def append_to_cgm_sheet(ws, paper: dict):
    """Add a new candidate row to the CGM table."""
    next_row = ws.max_row + 1
    values = [
        f"[NEW] {paper['title'][:60]}...",  # Dataset name placeholder
        paper["year"],
        "?",                                 # N — needs manual fill
        "? (see abstract)",                  # Population
        "? (see abstract)",                  # Dietary data
        "?",                                 # Meal times
        "?",                                 # CGM freq
        "?",                                 # File format
        "?",                                 # Duration
        f"PMID:{paper['pmid']}\ndoi:{paper['doi']}",
        f"{paper['authors']} {paper['journal']} {paper['year']}. Abstract: {paper['abstract'][:200]}...",
    ]
    for ci, val in enumerate(values, 1):
        ws.cell(row=next_row, column=ci, value=val)
    style_new_row(ws, next_row, len(values))
    ws.row_dimensions[next_row].height = 60


def append_to_omics_sheet(ws, paper: dict):
    """Add a new candidate row to the multi-omics table."""
    next_row = ws.max_row + 1
    values = [
        f"[NEW] {paper['title'][:60]}...",
        paper["year"],
        "?",
        "? (see abstract)",
        "? (see abstract)",
        "? (see abstract)",
        "?",
        "?",
        "?",
        f"PMID:{paper['pmid']}\ndoi:{paper['doi']}",
        f"{paper['authors']} {paper['journal']} {paper['year']}.",
    ]
    for ci, val in enumerate(values, 1):
        ws.cell(row=next_row, column=ci, value=val)
    style_new_row(ws, next_row, len(values))
    ws.row_dimensions[next_row].height = 60


# ─── MAIN ────────────────────────────────────────────────────────────────────

def run():
    print(f"\n{'='*60}")
    print(f"Literature Tracker — {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*60}\n")

    # Load spreadsheet
    xlsx_path = Path(SPREADSHEET_PATH)
    if not xlsx_path.exists():
        print(f"ERROR: Spreadsheet not found at {SPREADSHEET_PATH}")
        print("Please place your literature_tables.xlsx in the data/ folder.")
        return

    wb = openpyxl.load_workbook(xlsx_path)

    # Get the two sheets (by index in case names vary)
    sheet_names = wb.sheetnames
    print(f"Sheets found: {sheet_names}")
    ws_cgm   = wb[sheet_names[0]]  # Table 1 — CGM + Diet
    ws_omics = wb[sheet_names[1]]  # Table 2 — Multi-Omics

    # Load seen PMIDs (avoid duplicates across runs)
    seen_pmids = load_existing_pmids()
    new_pmids  = set()

    cgm_added   = 0
    omics_added = 0

    # ── Search for CGM + dietary papers ──
    print("Searching for CGM + dietary dataset papers...")
    for query in CGM_QUERIES:
        print(f"  Query: '{query}'")
        papers = search_pubmed(query, max_results=8)
        for paper in papers:
            if paper["pmid"] in seen_pmids:
                continue
            relevant, reason = is_relevant(paper, "cgm")
            if relevant:
                print(f"    ✓ NEW: {paper['title'][:70]}...")
                append_to_cgm_sheet(ws_cgm, paper)
                new_pmids.add(paper["pmid"])
                cgm_added += 1
            else:
                print(f"    ✗ Skipped ({reason}): {paper['title'][:50]}...")
                new_pmids.add(paper["pmid"])  # still mark as seen

    # ── Search for multi-omics papers ──
    print("\nSearching for multi-omics acute time series papers...")
    for query in OMICS_QUERIES:
        print(f"  Query: '{query}'")
        papers = search_pubmed(query, max_results=8)
        for paper in papers:
            if paper["pmid"] in seen_pmids or paper["pmid"] in new_pmids:
                continue
            relevant, reason = is_relevant(paper, "omics")
            if relevant:
                print(f"    ✓ NEW: {paper['title'][:70]}...")
                append_to_omics_sheet(ws_omics, paper)
                new_pmids.add(paper["pmid"])
                omics_added += 1
            else:
                print(f"    ✗ Skipped ({reason}): {paper['title'][:50]}...")
                new_pmids.add(paper["pmid"])

    # ── Add update timestamp to both sheets ──
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M UTC")
    for ws in [ws_cgm, ws_omics]:
        last_row = ws.max_row + 1
        ts_cell = ws.cell(row=last_row, column=1,
                          value=f"Last auto-updated: {timestamp} | New rows added this run: CGM={cgm_added}, Omics={omics_added}")
        ts_cell.font      = Font(italic=True, color="888888", name="Arial", size=8)
        ts_cell.alignment = Alignment(wrap_text=True)

    # ── Save ──
    wb.save(xlsx_path)
    print(f"\n{'='*60}")
    print(f"Done! Added {cgm_added} CGM rows, {omics_added} omics rows.")
    print(f"Spreadsheet saved: {xlsx_path}")
    print(f"{'='*60}\n")

    # Save log
    seen_pmids.update(new_pmids)
    save_seen_pmids(seen_pmids)

    # Write summary for GitHub Actions to pick up
    summary_path = Path("data/last_run_summary.json")
    with open(summary_path, "w") as f:
        json.dump({
            "run_at": timestamp,
            "cgm_rows_added": cgm_added,
            "omics_rows_added": omics_added,
            "total_new_pmids_seen": len(new_pmids),
        }, f, indent=2)


if __name__ == "__main__":
    run()
